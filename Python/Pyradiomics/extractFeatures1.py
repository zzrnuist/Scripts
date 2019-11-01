# coding=utf-8
from __future__ import print_function
import os
import six
from radiomics import featureextractor, setVerbosity
from openpyxl import Workbook

yamlDir = '/home/zzr/git/pyradiomics1111'
params = os.path.join(yamlDir, 'examples', 'exampleSettings', 'mine.yaml')
extractor = featureextractor.RadiomicsFeatureExtractor(params)
extractor.addProvenance(provenance_on=False)
extractor.enableAllFeatures()
# extractor.enableFeatureClassByName('shape', False)
# extractor.enableImageTypeByName('lbp', False)
# extractor.enableAllImageTypes()
# extractor.disableAllImageTypes()
# extractor.enableImageTypeByName('LoG', True)
setVerbosity(60)
file = Workbook()
table = file.create_sheet()
dataDir = '/media/zzr/My Passport/newMRI/IntensityStandardizationNii'
row = 1
for j, case in enumerate(sorted(os.listdir(dataDir))):
    if j >= 0:
        print(j, case)
        casePath = os.path.join(dataDir, case)
        img = [i for i in os.listdir(casePath) if 'img' in i and 'T2' in i]
        label = [i for i in os.listdir(casePath) if 'label' in i and 'T2' in i]
        for idxx, i in enumerate(img):
            imageName = os.path.join(casePath, i)
            maskName = os.path.join(casePath, label[idxx])
            result = extractor.execute(imageName, maskName, label=1)
            column = 1
            for key, val in six.iteritems(result):
                if row == 1:
                    table.cell(row=1, column=1, value='case')
                    table.cell(row=2, column=1, value=os.path.join(case, i))
                    table.cell(row=row, column=column+1, value=key)
                    table.cell(row=row + 1, column=column+1, value=str(val))
                else:
                    table.cell(row=row + 1, column=1, value=os.path.join(case, i))
                    table.cell(row=row + 1, column=column+1, value=str(val))
                column += 1

            row += 1
            assert len(result) == column-1

        file.save('/home/zzr/shared/T2_preprocess.xlsx')

